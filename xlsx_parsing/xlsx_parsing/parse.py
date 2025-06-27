from typing import Any, Dict, List, Union

import openpyxl

examples = [
    "example-excels/clearSKY data example 1 basic.xlsx",
    "example-excels/clearSKY data example 2 detailed.xlsx",
    "example-excels/clearSKY-question3.xlsx",
]

greek_mandatory_headers = [
    "Αριθμός Μητρώου",
    "Ονοματεπώνυμο",
    "Ακαδημαϊκό E-mail",
    "Περίοδος δήλωσης",
    "Τμήμα Τάξης",
    "Κλίμακα βαθμολόγησης",
    "Βαθμολογία",
]
english_mandatory_headers = {
    "student_id": "0000002",
    "name": "KARAGIANNHS ELENH",
    "course_id": "NTU_CS101",
    "exam_type": "Winter",
    "year": 2024,
    "grade": 6.0,
    "question_grades": [9.0, 3.0, 5.0, 7.0],
    "grade_weights": [1.0, 1.0, 1.0, 1.0],
}

greek_to_english = {
    "Αριθμός Μητρώου": "student_id",
    "Ονοματεπώνυμο": "name",
    "Βαθμολογία": "grade",
    "Περίοδος δήλωσης": "exam_type_and_year",
    "Κλίμακα βαθμολόγησης": "grade_scale",
    "Τμήμα Τάξης": "course_title",
    "Ακαδημαϊκό E-mail": "email",
}


def parse_grades_excel(file_path: str) -> Dict[str, Any]:
    result: Dict[str, Any] = {"error": "", "warning": "", "result": {}}

    try:
        # Load the workbook
        workbook = openpyxl.load_workbook(filename=file_path, data_only=True)
        sheet = workbook.active

        # Determine if it's simple or extended format
        is_extended = False
        num_questions = 0
        global_course_id = None
        global_exam_type = None
        global_year = None
        not_empty_rows = 0

        # Check for extended format markers
        if sheet.max_column >= 18:
            # Check for question headers in row 3, columns 9-18
            question_headers: List[str] = []
            for col in range(9, 19):
                cell_value = sheet.cell(row=3, column=col).value
                if cell_value and (
                    str(cell_value).startswith("Q") or str(cell_value).startswith("W")
                ):
                    question_headers.append(str(cell_value))

            if len(question_headers) > 0:
                is_extended = True
                num_questions = len(question_headers)

        # Validate basic structure
        if sheet.max_row < 4:
            result["error"] = "Excel file has too few rows (minimum 4 required)"
            return result

        # Extract headers (row 3)
        mandatory_headers: List[str] = []
        for col in range(1, 8):
            header = sheet.cell(row=3, column=col).value
            if header is None:
                result["error"] = f"Mandatory header in column {col} is missing"
                return result
            mandatory_headers.append(str(header))

        # Add question weights if extended format
        if is_extended:
            question_weights: List[float] = []
            for q in range(1, num_questions + 1):
                col = 8 + q
                weight = sheet.cell(row=2, column=col).value
                try:
                    question_weights.append(
                        float(weight) / 10 if weight is not None else 0.0
                    )  # adjustment (/10) so that weights sum up to 10
                except (ValueError, TypeError):
                    question_weights.append(0.0)
                    result[
                        "warning"
                    ] += f"Question weight in ({2},{col}) is not numeric, assumed as 0.0\n"

            if sum(question_weights) != 10:
                result[
                    "warning"
                ] += f"Question weights sum to {sum(question_weights)} instead of 10. Be sure that you have not forgotten to add headers (Q01, Q02, ...) for questions in the row 3 of the excel\n"

        # Process data rows
        data: List[Dict[str, Any]] = []
        for row in range(4, sheet.max_row + 1):
            student_data: Dict[str, Any] = {}

            if all((sheet.cell(row=row, column=c).value is None) for c in range(1, 8)):
                pass  # skip empty rows
            else:
                not_empty_rows += 1

            # Process mandatory columns (1-7)
            for col in range(1, 8):
                header = greek_to_english[mandatory_headers[col - 1]]
                if header in ["email", "grade_scale"]:
                    continue  # Skip headers that are not needed

                cell_value = sheet.cell(row=row, column=col).value
                if not cell_value and all(
                    (sheet.cell(row=row, column=c).value is None) for c in range(1, 8)
                ):
                    # Skip empty rows
                    continue
                elif not cell_value:
                    result["error"] = f"Mandatory data in ({row},{col}) is missing"
                    print("missed data")
                    return result

                if header == "exam_type_and_year":
                    # Split the exam type and year
                    try:
                        assert isinstance(cell_value, str)
                        year = int(cell_value.strip().split("-")[0])
                        exam_type = ""
                        if "ΧΕΙΜ" in cell_value:
                            exam_type = "Winter"
                        elif any(x in cell_value for x in ["ΕΑΡ", "ΘΕΡ"]):
                            exam_type = "Spring"
                        elif any(x in cell_value for x in ["ΣΕΠ", "ΕΠΑΝ"]):
                            exam_type = "September"
                        else:
                            raise ValueError("Invalid exam type")
                        student_data["exam_type"] = exam_type
                        student_data["year"] = year

                        # ensure no different exam types are present
                        if (not_empty_rows >= 2) and (global_exam_type != exam_type):
                            result["error"] = (
                                f"Exam type in ({row},{col}) does not match the global exam type: {global_exam_type}"
                            )
                            return result
                        global_exam_type = exam_type

                        # ensure no different years are present
                        if (not_empty_rows >= 2) and (global_year != year):
                            result["error"] = (
                                f"Year in ({row},{col}) does not match the global year: {global_year}"
                            )
                            return result
                        global_year = year

                    except Exception as e:
                        result["error"] = (
                            f"Invalid format in ({row},{col}) for exam type and year:\n {e}"
                        )
                        return result
                elif header == "course_title":
                    try:
                        assert isinstance(cell_value, str)
                        course_id = cell_value.strip().split("(")[1].split(")")[0]
                        student_data["course_id"] = course_id

                        # ensure no different course IDs are present
                        if (not_empty_rows >= 2) and (global_course_id != course_id):
                            result["error"] = (
                                f"Course ID in ({row},{col}) does not match the global course ID: {global_course_id}"
                            )
                            return result
                        global_course_id = course_id

                    except Exception as e:
                        result["error"] = (
                            f"Invalid format in ({row},{col}) for course title:\n {e}"
                        )
                        return result
                else:
                    student_data[header] = cell_value

            # Process question data if extended format
            if is_extended:
                question_scores: List[Union[int, float]] = []
                total_score = 0.0

                for q in range(1, num_questions + 1):
                    col = 8 + q
                    cell_value = sheet.cell(row=row, column=col).value

                    # Convert to float if possible, otherwise None
                    if isinstance(cell_value, (int, float)):
                        question_scores.append(cell_value)
                    elif isinstance(cell_value, str):
                        try:
                            question_scores.append(float(cell_value))
                        except ValueError:
                            question_scores.append(0.0)
                            result[
                                "warning"
                            ] += f"Question score in ({row},{col}) is not numeric, assumed as 0.0\n"
                            # result["error"] = f"Question score in ({row},{col}) is not numeric"
                            # return result
                    else:
                        question_scores.append(0.0)
                        result["warning"] = (
                            f"Question score in ({row},{col}) is not numeric, assumed as 0.0\n"
                        )
                        # result["error"] = f"Question score in ({row},{col}) is not numeric"
                        # return result

                # Calculate total score with None values treated as 0
                numeric_scores = [
                    question_scores[i]
                    * (
                        question_weights[i] / 10
                    )  # adjustment (/10) as weights sum up to 10
                    for i in range(len(question_scores))
                ]
                total_score = sum(numeric_scores) if numeric_scores else 0.0
                student_data["total_score"] = total_score

            if student_data:
                # add grade weights and question grades to student_data
                student_data["grade_weights"] = (
                    question_weights if is_extended else [10.0]
                )
                student_data["question_grades"] = (
                    question_scores if is_extended else [student_data["grade"]]
                )
                data.append(student_data)

        # Prepare result
        result["result"] = {
            "format": "extended" if is_extended else "simple",
            "num_questions": num_questions if is_extended else None,
            "mandatory_headers": mandatory_headers,
            "data": data,
            "question_weights": question_weights if is_extended else None,
        }

    except Exception as e:
        result["error"] = f"An error occurred while processing the file: {str(e)}"

    return result


def parse_enrolled_students_excel(file_path: str) -> Dict[str, Any]:
    result: Dict[str, Any] = {"error": "", "warning": "", "result": {}}
    student_ids: list[str] = []

    try:
        # Load the workbook
        workbook = openpyxl.load_workbook(filename=file_path, data_only=True)
        sheet = workbook.active

        global_course_id = None
        global_exam_type = None
        global_year = None
        not_empty_rows = 0

        # Validate basic structure
        if sheet.max_row < 4:
            result["error"] = "Excel file has too few rows (minimum 4 required)"
            return result

        # Extract headers (row 3)
        mandatory_headers: List[str] = []
        for col in range(
            1, 6
        ):  # student_id, name, email, exam_type_and_year, course_title
            header = sheet.cell(row=3, column=col).value
            if header is None:
                result["error"] = f"Mandatory header in column {col} is missing"
                return result
            mandatory_headers.append(str(header))

        # Process data rows
        for row in range(4, sheet.max_row + 1):
            student_data: Dict[str, Any] = {}

            if all((sheet.cell(row=row, column=c).value is None) for c in range(1, 8)):
                pass  # skip empty rows
            else:
                not_empty_rows += 1

            # Process mandatory columns (1-7)
            for col in range(1, 6):
                header = greek_to_english[mandatory_headers[col - 1]]
                if header in ["email", "grade_scale"]:
                    continue  # Skip headers that are not needed

                cell_value = sheet.cell(row=row, column=col).value
                if not cell_value and all(
                    (sheet.cell(row=row, column=c).value is None) for c in range(1, 6)
                ):
                    # Skip empty rows
                    continue
                elif not cell_value:
                    result["error"] = f"Mandatory data in ({row},{col}) is missing"
                    print("missed data")
                    return result

                if header == "exam_type_and_year":
                    # Split the exam type and year
                    try:
                        assert isinstance(cell_value, str)
                        year = int(cell_value.strip().split("-")[0])
                        exam_type = ""
                        if "ΧΕΙΜ" in cell_value:
                            exam_type = "Winter"
                        elif any(x in cell_value for x in ["ΕΑΡ", "ΘΕΡ"]):
                            exam_type = "Spring"
                        elif any(x in cell_value for x in ["ΣΕΠ", "ΕΠΑΝ"]):
                            exam_type = "September"
                        else:
                            raise ValueError("Invalid exam type")
                        student_data["exam_type"] = exam_type
                        student_data["year"] = year

                        # ensure no different exam types are present
                        if (not_empty_rows >= 2) and (global_exam_type != exam_type):
                            result["error"] = (
                                f"Exam type in ({row},{col}) does not match the global exam type: {global_exam_type}"
                            )
                            return result
                        global_exam_type = exam_type

                        # ensure no different years are present
                        if (not_empty_rows >= 2) and (global_year != year):
                            result["error"] = (
                                f"Year in ({row},{col}) does not match the global year: {global_year}"
                            )
                            return result
                        global_year = year

                    except Exception as e:
                        result["error"] = (
                            f"Invalid format in ({row},{col}) for exam type and year:\n {e}"
                        )
                        return result
                elif header == "course_title":
                    try:
                        assert isinstance(cell_value, str)
                        course_id = cell_value.strip().split("(")[1].split(")")[0]
                        student_data["course_id"] = course_id

                        # ensure no different course IDs are present
                        if (not_empty_rows >= 2) and (global_course_id != course_id):
                            result["error"] = (
                                f"Course ID in ({row},{col}) does not match the global course ID: {global_course_id}"
                            )
                            return result
                        global_course_id = course_id

                    except Exception as e:
                        result["error"] = (
                            f"Invalid format in ({row},{col}) for course title:\n {e}"
                        )
                        return result
                elif header == "student_id":
                    assert isinstance(cell_value, str)
                    student_ids.append(cell_value)
                    student_data[header] = cell_value

        # Prepare result
        result["result"] = {
            "format": "enrolled",
            "course_id": global_course_id,
            "exam_type": global_exam_type,
            "year": global_year,
            "student_ids": student_ids,
        }

    except Exception as e:
        result["error"] = f"An error occurred while processing the file: {str(e)}"

    return result


if __name__ == "__main__":
    file_path = examples[2]  # Example file path
    result = parse_grades_excel(file_path)

    print("Results:")
    if result["error"]:
        print(f"ERROR: {result['error']}")
    if result["warning"]:
        print(f"WARNING: {result['warning']}")
    if result["result"]:
        print("Parsed data structure:")
        print(f"Format: {result['result']['format']}")
        if result["result"]["format"] == "extended":
            print(f"Number of questions: {result['result']['num_questions']}")
            print(f"Question weights: {result['result']['question_weights']}")
        print(f"First student data: {result['result']['data'][0]}")
